#!/usr/bin/env python3
"""
Genera el SQL de carga inicial a partir del Excel "base de conocimiento".

    python3 scripts/generar-carga-excel.py <ruta-al-xlsx> > scripts/migrations/003_carga_excel.sql

Reglas de armado (ver el informe que imprime en stderr al final):
  · Negocio  = "NOMBRE CLIENTE" de la hoja de stock. Es el dueño del stock.
  · Sucursal = cada ficha de "Facturación e info clientes"; el negocio se
    detecta por prefijo (DESMADRE POCITOS → negocio DESMADRE, sucursal POCITOS).
  · Producto = nombre único del catálogo global; el precio y el stock viven en
    business_products, por negocio.
  · Precio vacío → NULL (a definir). Stock vacío → 0.
"""
import re
import sys
import unicodedata
import openpyxl

ALIAS = {"SUSSHI APP": "SUSHI APP"}   # typo en la hoja de facturación


def norm(s: str) -> str:
    s = unicodedata.normalize("NFD", (s or "").strip().upper())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s)
    for mal, bien in ALIAS.items():
        if s == mal or s.startswith(mal + " "):
            s = bien + s[len(mal):]
    return s


def numero(v):
    """Devuelve el valor numérico, o None si la celda trae texto."""
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def sql(v):
    if v is None or v == "":
        return "NULL"
    if isinstance(v, (int, float)):
        return repr(v)
    return "'" + str(v).replace("'", "''") + "'"


def partir_facturacion(texto):
    """Saca razón social, RUT y dirección del bloque de texto libre."""
    if not texto or texto.strip().lower().startswith("sin datos"):
        return None, None, None
    lineas = [l.strip() for l in str(texto).splitlines() if l.strip()]
    rut = None
    resto = []
    for l in lineas:
        m = re.search(r"RUT\s*[:\-]?\s*(\d[\d\.\s]{8,})", l, re.I)
        if m and not rut:
            rut = re.sub(r"\D", "", m.group(1))
            sobra = re.sub(r"RUT\s*[:\-]?\s*\d[\d\.\s]*", "", l, flags=re.I).strip(" -–,")
            if sobra:
                resto.append(sobra)
        else:
            resto.append(l)
    razon = resto[0] if resto else None
    direccion = "\n".join(resto[1:]) if len(resto) > 1 else None
    return razon, rut, direccion


def partir_info(texto):
    """Primera línea = dirección de entrega; lo que sigue a un guion, horario."""
    if not texto:
        return None, None
    lineas = [l.strip() for l in str(texto).splitlines() if l.strip()]
    if not lineas:
        return None, None
    primera = lineas[0]
    partes = re.split(r"\s+[–—-]\s+", primera, maxsplit=1)
    if len(partes) == 2 and re.search(r"\bhs\b|hora|despu[eé]s|antes|ma[ñn]ana|tarde", partes[1], re.I):
        return partes[0].strip(), partes[1].strip()
    return primera, None


def main(ruta):
    wb = openpyxl.load_workbook(ruta, data_only=True)

    # ── Hoja de stock ────────────────────────────────────────────────────────
    negocios = {}       # clave normalizada -> nombre visible
    productos = {}      # clave normalizada -> nombre visible
    lineas = []         # (negocio_key, producto_key, precio, stock, notas)
    descartes = []

    for nro, fila in enumerate(wb["Stock de productos"].iter_rows(min_row=2, values_only=True), start=2):
        cli, prod, precio, stock, notas = (list(fila) + [None] * 5)[:5]
        if not any([cli, prod, precio, stock, notas]):
            continue
        if cli and norm(cli).startswith("NOTA:"):
            descartes.append((nro, "nota suelta", str(cli)[:70]))
            continue
        if not cli:
            descartes.append((nro, "sin cliente", str(prod)[:70]))
            continue
        if not prod:
            descartes.append((nro, "sin producto", str(cli)[:70]))
            continue

        nk, pk = norm(cli), norm(prod)
        negocios.setdefault(nk, str(cli).strip())
        productos.setdefault(pk, str(prod).strip())

        # Precio y stock a veces traen texto ("revisar con felipe"): el valor
        # queda vacío y el texto se conserva como nota, no se descarta.
        aclaraciones = [str(notas).strip()] if notas else []
        precio_num = numero(precio)
        if precio not in (None, "") and precio_num is None:
            aclaraciones.append("Precio: %s" % str(precio).strip())
        stock_num = numero(stock)
        if stock not in (None, "") and stock_num is None:
            aclaraciones.append("Stock: %s" % str(stock).strip())

        lineas.append((nk, pk, precio_num,
                       int(stock_num) if stock_num is not None else 0,
                       " · ".join(aclaraciones) if aclaraciones else None))

    # ── Hoja de facturación ──────────────────────────────────────────────────
    fichas = []
    for fila in wb["Facturación e info clientes"].iter_rows(min_row=2, values_only=True):
        nombre, fact, info = (list(fila) + [None] * 3)[:3]
        if not nombre:
            continue
        fk = norm(nombre)
        # El negocio es el prefijo más largo que exista en la hoja de stock.
        cands = [n for n in negocios if fk == n or fk.startswith(n + " ")]
        if cands:
            neg = max(cands, key=len)
            suc = fk[len(neg):].strip() or None
        else:
            neg, suc = fk, None
            negocios.setdefault(fk, str(nombre).strip())
        fichas.append((neg, suc, fact, info))

    con_ficha = {n for n, _, _, _ in fichas}
    sin_ficha = [n for n in negocios if n not in con_ficha]

    # ── SQL ──────────────────────────────────────────────────────────────────
    out = sys.stdout
    out.write("-- Carga inicial generada por scripts/generar-carga-excel.py\n")
    out.write("-- Fuente: %s\n" % ruta.split("/")[-1])
    out.write("-- Correr DESPUÉS de 001_modelo_stock.sql\n\nBEGIN;\n\n")

    out.write("-- Negocios ------------------------------------------------------------\n")
    for nk in sorted(negocios):
        out.write("INSERT INTO businesses (nombre) VALUES (%s);\n" % sql(negocios[nk]))

    out.write("\n-- Sucursales (facturación y entrega) ---------------------------------\n")
    for neg, suc, fact, info in fichas:
        razon, rut, dir_fact = partir_facturacion(fact)
        dir_ent, horario = partir_info(info)
        out.write(
            "INSERT INTO clients (business_id, sucursal, razon_social, rut, "
            "direccion_facturacion, direccion_entrega, horario_entrega, "
            "datos_facturacion, info_cliente)\n"
            "  SELECT id, %s, %s, %s, %s, %s, %s, %s, %s FROM businesses WHERE nombre = %s;\n"
            % (sql(suc), sql(razon), sql(rut), sql(dir_fact), sql(dir_ent), sql(horario),
               sql(fact), sql(info), sql(negocios[neg]))
        )
    for nk in sorted(sin_ficha):
        out.write(
            "INSERT INTO clients (business_id) SELECT id FROM businesses WHERE nombre = %s;\n"
            % sql(negocios[nk])
        )

    out.write("\n-- Catálogo de productos ----------------------------------------------\n")
    for i, pk in enumerate(sorted(productos), start=1):
        out.write("INSERT INTO products (codigo_prod, nombre) VALUES ('P-%04d', %s);\n"
                  % (i, sql(productos[pk])))

    out.write("\n-- Precio y stock por negocio -----------------------------------------\n")
    for nk, pk, precio, stock, notas in lineas:
        out.write(
            "INSERT INTO business_products (business_id, product_id, precio, stock, notas)\n"
            "  SELECT b.id, p.id, %s, %s, %s FROM businesses b, products p\n"
            "   WHERE b.nombre = %s AND p.nombre = %s;\n"
            % (sql(precio), stock, sql(notas), sql(negocios[nk]), sql(productos[pk]))
        )

    out.write("\nCOMMIT;\n")

    # ── Informe ──────────────────────────────────────────────────────────────
    e = sys.stderr
    e.write("\n=== RESUMEN DE LA CARGA ===\n")
    e.write("  negocios .......... %d\n" % len(negocios))
    e.write("  sucursales ........ %d (%d de fichas + %d creadas vacías)\n"
            % (len(fichas) + len(sin_ficha), len(fichas), len(sin_ficha)))
    e.write("  productos ......... %d\n" % len(productos))
    e.write("  precio+stock ...... %d filas\n" % len(lineas))
    e.write("  sin precio ........ %d\n" % sum(1 for l in lineas if l[2] is None))
    e.write("  stock en cero ..... %d\n" % sum(1 for l in lineas if l[3] == 0))
    if sin_ficha:
        e.write("\n  Negocios sin datos de facturación (%d): %s\n"
                % (len(sin_ficha), ", ".join(sorted(negocios[n] for n in sin_ficha))))
    if descartes:
        e.write("\n  Filas NO cargadas (%d):\n" % len(descartes))
        for nro, motivo, txt in descartes:
            e.write("    fila %d — %s: %s\n" % (nro, motivo, txt))


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit("uso: generar-carga-excel.py <archivo.xlsx>")
    main(sys.argv[1])
